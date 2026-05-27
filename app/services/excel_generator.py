import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelGenerator:
    @staticmethod
    def generate_transactions_excel(
        resumen_info: dict,
        transacciones: list,
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transacciones"

        ws.merge_cells("A1:G1")
        cell = ws.cell(
            row=1,
            column=1,
            value=f"Transacciones Extraidas - {resumen_info.get('periodo', '')}",
        )
        cell.font = Font(bold=True, size=14)

        ws.merge_cells("A2:G2")
        ws.cell(
            row=2, column=1, value=f"Tarjeta: {resumen_info.get('tipo', '')} - {len(transacciones)} transacciones"
        )

        headers = ["Fecha", "Descripcion", "Monto", "Tarjeta", "Moneda", "Cuotas", "Cuota N°"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="2563EB", end_color="2563EB", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(transacciones, 5):
            ws.cell(row=i, column=1, value=t.get("fecha", ""))
            ws.cell(row=i, column=2, value=t.get("descripcion", ""))
            ws.cell(row=i, column=3, value=t.get("monto", 0))
            ws.cell(row=i, column=4, value=t.get("numero_tarjeta", ""))
            ws.cell(row=i, column=5, value=t.get("moneda", "ARS"))
            ccs = t.get("cantidad_cuotas") or 1
            ws.cell(row=i, column=6, value=ccs if ccs > 1 else "-")
            ws.cell(row=i, column=7, value=t.get("cuota_numero", "-") if ccs > 1 else "-")

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_unmatched_excel(
        resumen_info: dict,
        transacciones: list,
        nombre_archivo: str = "pagos_sin_factura.xlsx",
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos sin Factura"

        ws.merge_cells("A1:F1")
        cell = ws.cell(
            row=1,
            column=1,
            value=f"Pagos sin Factura Asociada - {resumen_info.get('periodo', '')}",
        )
        cell.font = Font(bold=True, size=14)

        ws.merge_cells("A2:F2")
        ws.cell(
            row=2, column=1, value=f"Tipo: {resumen_info.get('tipo', '')}"
        )

        headers = ["Fecha", "Descripción", "Monto", "Tarjeta", "Moneda", "Período"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="2563EB", end_color="2563EB", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(transacciones, 5):
            ws.cell(row=i, column=1, value=t.get("fecha", ""))
            ws.cell(row=i, column=2, value=t.get("descripcion", ""))
            ws.cell(row=i, column=3, value=t.get("monto", 0))
            ws.cell(row=i, column=4, value=t.get("numero_tarjeta", ""))
            ws.cell(row=i, column=5, value=t.get("moneda", "ARS"))
            ws.cell(row=i, column=6, value=resumen_info.get("periodo", ""))

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_invoices_preview_excel(facturas: list) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vista previa"

        headers = [
            "Archivo", "Método", "Emisor", "CUIT", "Fecha", "Vencimiento",
            "Monto", "Subtotal", "Moneda", "Tipo", "N° Factura", "Cuota N°",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="2563EB", end_color="2563EB", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for i, f in enumerate(facturas, 2):
            datos = f.get("datos") or {}
            ws.cell(row=i, column=1, value=f.get("drive_file_name", ""))
            ws.cell(row=i, column=2, value=f.get("extraction_method", ""))
            ws.cell(row=i, column=3, value=datos.get("emisor", ""))
            ws.cell(row=i, column=4, value=datos.get("cuit_emisor", ""))
            ws.cell(row=i, column=5, value=datos.get("fecha", ""))
            ws.cell(row=i, column=6, value=datos.get("vencimiento", ""))
            ws.cell(row=i, column=7, value=datos.get("monto_total", ""))
            ws.cell(row=i, column=8, value=datos.get("subtotal", ""))
            ws.cell(row=i, column=9, value=datos.get("moneda", ""))
            ws.cell(row=i, column=10, value=datos.get("tipo_factura", ""))
            ws.cell(row=i, column=11, value=datos.get("numero_factura", ""))
            cn = datos.get("cuota_numero")
            ws.cell(row=i, column=12, value=cn if cn is not None else "")

        col_letters = [chr(65 + i) for i in range(12)]
        widths = [30, 16, 24, 18, 14, 14, 14, 14, 10, 10, 16, 10]
        for letter, w in zip(col_letters, widths):
            ws.column_dimensions[letter].width = w

        ws2 = wb.create_sheet("Texto extraído")
        ws2.cell(row=1, column=1, value="Archivo").font = Font(bold=True)
        ws2.cell(row=1, column=2, value="Texto").font = Font(bold=True)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 100

        for i, f in enumerate(facturas, 2):
            ws2.cell(row=i, column=1, value=f.get("drive_file_name", ""))
            ws2.cell(row=i, column=2, value=f.get("raw_text", ""))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
